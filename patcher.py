# patcher.py
# 一个专门用于读取现有Excel，并为缺失截图的订单重新截图并更新文件的工具

import time
import os
import io

from PIL import Image, ImageChops
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

# ==================== 用户配置区 ====================
# 1. 您那个半成品的Excel文件名
EXCEL_FILE_PATH = "微店订单导出(copy).xlsx"

# 2. 订单号所在的列名 (根据您的截图是N列)
ORDER_ID_COLUMN_NAME = "微店/淘宝订单号"

# 3. 截图要写入的列 (根据您的截图是P列)
SCREENSHOT_COLUMN_LETTER = "P"

# 4. 从Excel的第几行开始检查 (2336行，忽略表头)
START_ROW = 3576

MAX_RUNS = 600

# 5. 在这里粘贴您最新的有效Cookie
MY_COOKIE = ("duid=1914883825; is_login=true; login_source=LOGIN_USER_SOURCE_MASTER; login_token=_EwWqqVIQBKfktIYZo22vw1TIKUisg0BeMMyiHt9lfxnAv1_0ILkdXIyLAcmu6HK0hlNUYYWc68AOePlbZGyK8z9CHVIjMgPEwLeWRGIDQBoSOzm9THeoRdEFoPTMjGrgLuGYX8yAvOih15k2YY1u3sD-W84621RdhI9TP9h_YYSGZBom9XRr07xvXqVgBzD9wrtS97w0Ew2mivMx1UYUVX9m7xG8FFCphKEkrEWEWXQ0j37JqFOoB28pRJQxYmD7zU2YnB-E; login_type=LOGIN_USER_TYPE_MASTER; sid=1798256885; smart_login_type=0; uid=1914883825; __spider__sessionid=b9ff407a1e57023d; __spider__visitorid=10bfef0417a79d8f; is_follow_mp=0; wdtoken=00d03344")


# ====================================================


class Patcher:
    """
    一个专门修复Excel中缺失截图的工具类。
    """

    def __init__(self, headless: bool = True):
        print("正在初始化Selenium浏览器驱动...")
        options = webdriver.ChromeOptions()
        options.add_experimental_option('excludeSwitches', ['enable-automation'])
        options.add_argument("--window-size=1200,1080")
        if headless:
            options.add_argument("--headless")
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=options)
        self.wait = WebDriverWait(self.driver, 30)
        print("浏览器驱动初始化完成。")

    def login_with_cookie(self, cookie_string: str):
        print("正在使用Cookie登录...")
        self.driver.get("https://weidian.com/")
        time.sleep(2)
        self.driver.delete_all_cookies()
        for cookie_item in cookie_string.split(';'):
            if '=' in cookie_item:
                name, value = cookie_item.strip().split('=', 1)
                self.driver.add_cookie({'name': name, 'value': value, 'domain': '.weidian.com'})
        print("Cookie注入完成。")

    def _trim_image(self, image: Image.Image) -> Image.Image:
        bg = Image.new(image.mode, image.size, image.getpixel((0, 0)))
        diff = ImageChops.difference(image, bg)
        diff = ImageChops.add(diff, diff, 2.0, -100)
        bbox = diff.getbbox()
        if bbox:
            return image.crop(bbox)
        return image

    def take_single_screenshot(self, order_id: str) -> str:
        """
        根据单个订单ID，访问详情页并返回截图的本地路径。
        """
        if not order_id:
            raise ValueError("订单ID为空，无法截图。")

        # 拼接详情页URL
        detail_url = f"https://i.weidian.com/order/detail.php?oid={order_id}"

        self.driver.get(detail_url)
        main_container = self.wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="detail"]')))
        self.driver.execute_script("document.body.style.zoom='40%'")
        time.sleep(1.5)

        base_screenshot_png = main_container.screenshot_as_png
        base_image = Image.open(io.BytesIO(base_screenshot_png))

        final_image = base_image.crop((0, 0, 640, base_image.height))
        final_image = self._trim_image(final_image)

        if not os.path.exists("screenshots_patch"):
            os.makedirs("screenshots_patch")

        screenshot_path = os.path.join("screenshots_patch", f"{order_id}.png")
        final_image.save(screenshot_path)
        return screenshot_path

    def run(self):
        # 0. 【核心修改】检查文件是否存在
        if not os.path.exists(EXCEL_FILE_PATH):
            print(f"错误：找不到指定的Excel文件 '{EXCEL_FILE_PATH}'。请确保文件名正确且文件在脚本同目录下。")
            return

        # 1. 登录
        self.login_with_cookie(MY_COOKIE)

        # 2. 加载现有的Excel工作簿和工作表
        print(f"正在加载Excel文件: {EXCEL_FILE_PATH}...")
        # 使用 openpyxl 直接加载，准备进行原地编辑
        workbook = load_workbook(EXCEL_FILE_PATH)
        sheet = workbook.active
        print("Excel文件加载成功。")

        # 3. 找到订单号列的索引
        headers = [cell.value for cell in sheet[1]]
        try:
            # 找到订单号所在的列的索引 (注意，openpyxl的列索引从1开始)
            order_id_col_index = headers.index(ORDER_ID_COLUMN_NAME) + 1
        except ValueError:
            print(f"错误：在Excel的表头中未找到订单号列 '{ORDER_ID_COLUMN_NAME}'。")
            return

        screenshot_col_letter = SCREENSHOT_COLUMN_LETTER

        # 4. 遍历指定的行范围，并加入MAX_RUNS计数器
        print(f"将从第 {START_ROW} 行开始处理，最多处理 {MAX_RUNS} 个订单...")

        runs_count = 0
        for row_num in range(START_ROW, sheet.max_row + 1):
            if 'MAX_RUNS' in globals() and runs_count >= MAX_RUNS:
                print(f"\n已达到本次试运行的最大数量 ({MAX_RUNS} 个)，提前结束。")
                break

            # 使用列索引来获取单元格，更稳健
            order_id = sheet.cell(row=row_num, column=order_id_col_index).value

            if not order_id or not isinstance(order_id, (str, int)):
                print(f"  > 第 {row_num} 行订单号为空或格式无效，跳过。")
                continue
            order_id = str(order_id)

            print(
                f"--- 正在处理第 {row_num} 行, 订单ID: {order_id} ({runs_count + 1}/{MAX_RUNS if 'MAX_RUNS' in globals() else '无限'}) ---")

            try:
                # 5. 执行截图
                screenshot_path = self.take_single_screenshot(order_id)

                # 6. 【核心修改】将新截图“添加”到工作表中
                img = OpenpyxlImage(screenshot_path)

                # 设置图片的锚点，将其“钉”在目标单元格上
                img.anchor = f"{screenshot_col_letter}{row_num}"
                sheet.add_image(img)

                print(f"  > ✅ 成功为订单 {order_id} 在原文件上添加新截图。")

            except Exception as e:
                print(f"  > ❌ 处理订单 {order_id} 时发生错误: {e}")
                continue

            runs_count += 1

        # 7. 【核心修改】直接保存对原始文件的修改
        print(f"\n所有订单处理完毕，正在保存对文件 '{EXCEL_FILE_PATH}' 的修改...")
        workbook.save(EXCEL_FILE_PATH)
        print("🎉 补丁任务完成！文件已原地更新。")


if __name__ == "__main__":
    patcher = Patcher(headless=False)  # 建议首次运行时设置为False，方便观察
    try:
        if "YOUR_LATEST_COOKIE_HERE" in MY_COOKIE:
            raise ValueError("请在脚本顶部的配置区填写您最新的有效Cookie！")
        patcher.run()
    except Exception as e:
        print(f"程序执行时遇到致命错误: {e}")
    finally:
        print("程序运行结束，按回车键退出...")
        input()
        patcher.driver.quit()